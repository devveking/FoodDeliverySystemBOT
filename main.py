
TOKEN = 'your token'

import logging
import openpyxl
from aiogram import Bot, Dispatcher, types



bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
logging.basicConfig(level=logging.INFO)


orders = {}


menu = {
    "Здоровая еда": [
        {"name": "Овсянка с ягодами", "price": 690, "photo_url": "https://img.freepik.com/premium-photo/oatmeal-with-berries-and-fruits_73872-1027.jpg"},
        {"name": "Авокадо-тост", "price": 500, "photo_url": "https://as1.ftcdn.net/v2/jpg/01/44/48/96/1000_F_144489628_G6KKcaz86g0Jrw71HKy3VRUMGbkEWfsd.jpg"},
        {"name": "Cалат с киноа", "price": 300, "photo_url": "https://img.iamcook.ru/2022/upl/recipes/cat/u-ba9c5ab9f820c1a882343092421e49f4.jpg"},
    ],
    "Свежевыжатые соки": [
        {"name": "Смузи с киви", "price": 200, "photo_url": "https://i.obozrevatel.com/food/recipemain/2019/3/18/ss.jpg?size=636x424"},
        {"name": "Фреш из апельсина", "price": 150, "photo_url": "https://i.obozrevatel.com/food/recipemain/2019/9/11/apelsinovyj-sok5-1.jpg?size=636x424"},
        {"name": "Гранатовый сок", "price": 200, "photo_url": "https://media-cdn.tripadvisor.com/media/photo-s/1b/a4/77/54/caption.jpg"},
    ],
}


@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    user_id = message.from_user.id
    orders[user_id] = {'order_items': []}
    await set_persistent_menu(user_id)
    await message.answer("Привет! Я бот для заказа еды в ваш университет. Выберите категорию:")


async def set_persistent_menu(user_id):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, selective=True)
    keyboard.add("Начать")
    keyboard.add("Здоровая еда", "Свежевыжатые соки")
    keyboard.add("Корзина", "Оформить заказ")
    await bot.send_message(user_id, "Выберите категорию блюда:", reply_markup=keyboard)

@dp.message_handler(lambda message: message.text == "Начать")
async def start_over(message: types.Message):
    user_id = message.from_user.id
    orders[user_id] = {'order_items': []}
    await set_persistent_menu(user_id)
    await message.answer("Привет! Я бот для заказа еды в ваш университет. Выберите категорию:")



@dp.message_handler(lambda message: message.text in menu.keys())
async def choose_category(message: types.Message):
    user_id = message.from_user.id
    category = message.text
    orders[user_id]['current_category'] = category
    category_items = menu[category]
    for item in category_items:
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton(text=f"Добавить в корзину ({item['price']} tg.)", callback_data=f"add_to_cart {item['name']}"))
        await bot.send_photo(message.from_user.id, item['photo_url'], caption=f"{item['name']} ({item['price']} tg.)", reply_markup=keyboard)


@dp.callback_query_handler(lambda query: query.data.startswith("add_to_cart"))
async def add_to_cart(query: types.CallbackQuery):
    user_id = query.from_user.id
    category = orders[user_id]['current_category']
    item_name = query.data.split(" ")[1]
    item = next((x for x in menu[category] if x["name"] == item_name), None)
    if item:
        orders[user_id]['order_items'].append(item)
        await bot.answer_callback_query(query.id, text=f"{item_name} добавлено в корзину.")


@dp.message_handler(lambda message: message.text == "Корзина")
async def show_cart(message: types.Message):
    user_id = message.from_user.id
    if not orders[user_id]['order_items']:
        await message.answer("Ваша корзина пуста.")
    else:
        total_price = sum(item['price'] for item in orders[user_id]['order_items'])
        cart_text = "Ваша корзина:\n\n"
        for item in orders[user_id]['order_items']:
            cart_text += f"{item['name']} - {item['price']} tg.\n"
        cart_text += f"Итого: {total_price} tg."
        await message.answer(cart_text)


@dp.message_handler(lambda message: message.text == "Оформить заказ")
async def checkout(message: types.Message):
    user_id = message.from_user.id
    if not orders[user_id]['order_items']:
        await message.answer("Ваша корзина пуста. Нельзя оформить заказ.")
        return
    keyboard = types.ReplyKeyboardRemove()
    await message.answer("Введите ваше имя:")
    orders[user_id]['checkout'] = True


@dp.message_handler(lambda message: 'checkout' in orders.get(message.from_user.id, {}))
async def process_checkout(message: types.Message):
    user_id = message.from_user.id
    if 'name' not in orders[user_id]:
        orders[user_id]['name'] = message.text
        await message.answer("Введите ваш номер телефона:")
    else:
        orders[user_id]['phone'] = message.text
        save_order_to_excel(user_id)
        del orders[user_id]
        await message.answer("Спасибо за заказ! Ваш заказ оформлен. Если хотите заказать еще нажмите на кнопку 'Начать' ")


def save_order_to_excel(user_id):
    wb = openpyxl.load_workbook('/Users/anuar/PycharmProjects/pythonBOKZHE/sdufood.xlsx')
    sheet = wb.active
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=user_id)
    sheet.cell(row=next_row, column=2, value=orders[user_id]['name'])
    sheet.cell(row=next_row, column=3, value=orders[user_id]['phone'])
    for item in orders[user_id]['order_items']:
        sheet.cell(row=next_row, column=4, value=item['name'])
        sheet.cell(row=next_row, column=5, value=item['price'])
        next_row += 1
    wb.save('/Users/anuar/PycharmProjects/pythonBOKZHE/sdufood.xlsx')


from aiogram import executor
executor.start_polling(dp, skip_updates=True)
